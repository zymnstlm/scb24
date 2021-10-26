# Author : 柠檬班-亚萌
# Project : scb24
# Time : 2021/10/21 21:47
# E-mail : 3343787213@qq.com
# Company : 湖南零檬信息技术有限公司
# Site : http://www.lemonban.com
# Forum : http://testingpai.com


# 自动化测试流程
'''
1、设计自动化用例，通过自动化读取excel里面的数据  -- read_data()
2、编写自动化代码，发送请求   -- smarthome_post()
3、预期结果  vs  实际结果
4、再把最终结果写回到excel    -- write_data()
5、发送测试报告
'''

import openpyxl
import requests
def read_data(filename,sheet):
    workbook = openpyxl.load_workbook(filename)  #　加载工作簿到python内存
    sheet = workbook[sheet]  # 读取表单
    # max_row#获取总行数
    max_row = sheet.max_row  # 设置最大行数变量
    list_1 = []  # 定义一个空列表，用来装测试数据
    for i in range(2,max_row+1,1):  # 通过最大行数来遍历整个表单里的测试数据
        dict_1 = dict(
        case_id = sheet.cell(row=i, column=1).value,  # 获取用例编号
        header = sheet.cell(row=i, column=5).value,  # 获取请求头
        url = sheet.cell(row=i, column=6).value,  # 获取接口地址
        data = sheet.cell(row=i, column=7).value,  # 获取请求参数
        expected = sheet.cell(row=i, column=8).value  # 获取预期结果
        )
        list_1.append(dict_1)  # 将测试数据一条一条的追加到列表中保存
        # print(case_id,header_login,url_login,data_login)
        # print(dict_1)
    # print(list_1)
    return list_1
def smarthome_post(url,body,header):
    res = requests.post(url=url,json=body,headers=header)
    response = res.json()
    return response
def write_data(filename,sheet,row,final_result):
    workbook = openpyxl.load_workbook(filename)  #　加载工作簿到python内存
    sheet = workbook[sheet]  # 读取表单
    sheet.cell(row=row, column=9).value = final_result  # 写入最终结果
    workbook.save(filename)


def exe_test(filename,sheetname):
    cases = read_data(filename, sheetname)
    for case in cases:
        case_id = case['case_id']  # 取出用例编号
        url = case['url']  # 取出接口地址
        data = case['data']  # 取出请求参数
        # eval()：运行被字符串包裹着的python表达式 -- 相当于去掉字符串的单引号
        data = eval(data)  # 使用eval函数，将字符串转换成字典
        header = case['header']  # 取出请求头
        header = eval(header)  #使用eval函数，将字符串转换成字典、
        expected = case['expected']  # 取出预期结果
        expected = eval(expected)  # 使用eval函数，将字符串转换成字典
        expected_code = expected['code']  # 取出预期结果里面的code
        real_result = smarthome_post(url=url,body=data,header=header)  # 调用发送请求函数，进行接口的执行
        real_result_code = real_result['code'] # 取出实际结果里面的code
        real_result_code = eval(real_result_code)  # 使用eval函数，将字符串转换成字典
        print('预期结果code为：{}'.format(expected_code))
        print('实际结果code为：{}'.format(real_result_code))
        if expected_code == real_result_code:   #  断言，做结果的判断
            # if expected_msg == real_result_msg:
            print("{}模块第{}条用例是通过的！！".format(sheetname,case_id))
            print("*" * 40)
            final_result = 'pass'
        else:
            print("{}模块第{}条用例不通过！！".format(sheetname,case_id))
            print("*" * 40)
            final_result = 'fail'
        write_data(filename,sheetname,case_id+1,final_result)


exe_test('testcase_api_wuye.xlsx', 'login')
exe_test('testcase_api_wuye.xlsx', 'register')